from django.contrib import messages
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.http import HttpResponse
import re
import json
import zipfile
import io
import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from .forms import DocumentUploadForm, DocumentCommentForm, DocumentCorrectionForm, UserRegistrationForm
from .models import Document, DocumentHistory, ERPExport, AuditLog, Profile
from .utils import extract_text_from_pdf, send_rejection_email, send_approval_email, send_validator_notification
from .document_validator import DocumentValidator, FaceAnalyzer


def is_worker(user):
    return hasattr(user, "profile") and user.profile.is_worker


def is_admin(user):
    return hasattr(user, "profile") and user.profile.role == "admin"


def validate_curp(curp):
    """Valida formato estándar CURP mexicano (flexible)"""
    if not curp:
        return False
    
    curp = str(curp).strip().upper()
    
    # CURP válido tiene 17 o 18 caracteres
    if len(curp) not in (17, 18):
        return False
    
    # Formato flexible: 
    # 4 letras + 6 dígitos (YYMMDD) + 1 letra (H/M) + 6-7 caracteres alfanuméricos
    if len(curp) == 18:
        pattern = r"^[A-Z]{4}\d{6}[HM][A-Z0-9]{8}$"
    else:  # 17 caracteres
        pattern = r"^[A-Z]{4}\d{6}[HM][A-Z0-9]{7}$"
    
    return re.match(pattern, curp) is not None


def validate_rfc(rfc):
    """Valida formato estándar RFC mexicano"""
    if not rfc:
        return False
    
    rfc = str(rfc).strip().upper()
    
    # RFC válido tiene 12 o 13 caracteres
    # Formato: 3-4 letras + 6 dígitos (YYMMDD) + 6 caracteres (3 letras + 3 dígitos/letras)
    if len(rfc) < 12 or len(rfc) > 13:
        return False
    
    pattern = r"^[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{6}$"
    return re.match(pattern, rfc) is not None


def register(request):
    if request.method == "POST":
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Registro completo. Ya puedes subir documentos.")
            return redirect("home")
    else:
        form = UserRegistrationForm()

    return render(request, "documents/register.html", {"form": form})


@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "Has cerrado sesión exitosamente.")
    return redirect("home")


@login_required
def change_password_required(request):
    """Vista obligatoria para cambiar contraseña temporal (TempPass123!)"""
    # Verificar si la contraseña es la temporal
    user = request.user
    is_temp_password = user.check_password("TempPass123!")
    
    if not is_temp_password:
        # Si no tiene contraseña temporal, redirigir a home
        return redirect("home")
    
    if request.method == "POST":
        current_password = request.POST.get("current_password", "")
        new_password = request.POST.get("new_password", "")
        confirm_password = request.POST.get("confirm_password", "")
        
        # Validar contraseña actual
        if not user.check_password(current_password):
            messages.error(request, "❌ Contraseña actual incorrecta.")
            return render(request, "documents/change_password_required.html", {"temp_user": True})
        
        # Validar que coincidan
        if new_password != confirm_password:
            messages.error(request, "❌ Las contraseñas no coinciden.")
            return render(request, "documents/change_password_required.html", {"temp_user": True})
        
        # Validar longitud mínima
        if len(new_password) < 8:
            messages.error(request, "❌ La contraseña debe tener al menos 8 caracteres.")
            return render(request, "documents/change_password_required.html", {"temp_user": True})
        
        # Cambiar contraseña
        user.set_password(new_password)
        user.save()
        update_session_auth_hash(request, user)  # Mantener la sesión activa
        
        messages.success(request, "✅ Contraseña cambiada correctamente. Ya puedes acceder al sistema.")
        
        # Redirigir según el rol
        if hasattr(user, 'profile'):
            if user.profile.role == 'admin':
                return redirect('admin_panel')
            elif user.profile.role == 'validator1':
                return redirect('validator1_panel')
            elif user.profile.role == 'validator2':
                return redirect('validator2_panel')
        
        return redirect("home")
    
    return render(request, "documents/change_password_required.html", {"temp_user": True})


@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "Has cerrado sesión exitosamente.")
    return redirect("home")


def add_history(document, status, note="", author=None):
    DocumentHistory.objects.create(
        document=document,
        status=status,
        note=note,
        author=author,
    )


def determine_client_type(document) -> dict:
    """
    Determina el tipo de cliente y su estado:
    - 'new': Cliente nuevo (primera vez)
    - 'old_match': Cliente antiguo con datos que matchean
    - 'old_no_match': Cliente antiguo con datos que NO matchean
    - 'suspicious': Cliente con posible fraude
    """
    from django.contrib.auth.models import User
    from datetime import timedelta
    
    user = document.user
    now = timezone.now()
    
    # Contar documentos previos del usuario
    previous_docs = Document.objects.filter(
        user=user, 
        uploaded_at__lt=document.uploaded_at
    ).count()
    
    # Si es el primer documento, es cliente nuevo
    if previous_docs == 0:
        return {
            'type': 'new',
            'label': '🆕 CLIENTE NUEVO',
            'alert': 'Este es el primer documento del cliente en el sistema.',
            'color': '#3b82f6',  # azul
            'recommendation': 'Verificar datos personales con cuidado'
        }
    
    # Si hay documentos previos, es cliente antiguo
    # Verificar si la información matchea
    consistent = True
    issues = []
    
    if document.curp and previous_docs > 0:
        prev_doc = Document.objects.filter(user=user, curp__isnull=False).first()
        if prev_doc and prev_doc.curp != document.curp:
            consistent = False
            issues.append("CURP no coincide con registros anteriores")
    
    if document.rfc and previous_docs > 0:
        prev_doc = Document.objects.filter(user=user, rfc__isnull=False).first()
        if prev_doc and prev_doc.rfc != document.rfc:
            consistent = False
            issues.append("RFC no coincide con registros anteriores")
    
    if consistent:
        return {
            'type': 'old_match',
            'label': '✓ CLIENTE ANTIGUO (Datos Consistentes)',
            'alert': f'Cliente con {previous_docs} solicitud(es) previa(s). Datos coinciden.',
            'color': '#16a34a',  # verde
            'recommendation': 'Proceso normal, datos verificados'
        }
    else:
        return {
            'type': 'old_no_match',
            'label': '⚠️ CLIENTE ANTIGUO (Datos Inconsistentes)',
            'alert': f'Cliente con {previous_docs} solicitud(es) previa(s) pero DATOS NO COINCIDEN',
            'issues': issues,
            'color': '#dc2626',  # rojo
            'recommendation': 'REVISAR CUIDADOSAMENTE - Posible error de datos o fraude'
        }


def perform_ai_validation(document) -> dict:
    """
    Realiza validación automática inteligente del documento usando OCR y validaciones
    
    Returns: {
        'success': bool,
        'ocr_confidence': float,
        'overall_score': float,
        'recommendation': str,
        'curp_valid': bool,
        'rfc_valid': bool,
        'data_consistent': bool,
        'details': dict,
        'can_auto_approve': bool
    }
    """
    try:
        validator = DocumentValidator()
        
        files_to_validate = []
        if document.ine:
            files_to_validate.append(document.ine)
        if document.acta_nacimiento:
            files_to_validate.append(document.acta_nacimiento)
        if document.constancia_fiscal:
            files_to_validate.append(document.constancia_fiscal)
        if document.curp_documento:
            files_to_validate.append(document.curp_documento)
        if document.file:
            files_to_validate.append(document.file)
        
        if not files_to_validate:
            return {
                'success': False,
                'error': 'No hay archivo para validar',
                'can_auto_approve': False
            }
        
        best_result = None
        best_score = 0
        
        for f in files_to_validate:
            ai_result = validator.extract_and_validate(f.path)
            score = ai_result.get('overall_score', 0)
            if score > best_score:
                best_score = score
                best_result = ai_result
        
        if not best_result:
            return {
                'success': False,
                'error': 'No se pudo extraer texto de ningun documento',
                'can_auto_approve': False
            }
        
        ai_result = best_result
        
        # Parsea resultados
        curp_valid = ai_result.get('curp_validation', {}).get('valid', False) if ai_result.get('curp_validation') else False
        rfc_valid = ai_result.get('rfc_validation', {}).get('valid', False) if ai_result.get('rfc_validation') else False
        data_consistent = ai_result.get('data_consistency', {}).get('consistent', True) if ai_result.get('data_consistency') else False
        
        overall_score = ai_result.get('overall_score', 0)
        recommendation = ai_result.get('recommendation', 'REVISAR_MANUALMENTE').replace('_', ' ')
        
        # Determine si puede ser auto-aprobado
        can_auto_approve = (
            recommendation == 'VALIDAR_AUTOMATICAMENTE' and
            curp_valid and
            rfc_valid and
            data_consistent and
            overall_score >= 85
        )
        
        return {
            'success': True,
            'ocr_confidence': ai_result.get('ocr', {}).get('confidence', 0),
            'overall_score': overall_score,
            'recommendation': recommendation,
            'curp_valid': curp_valid,
            'rfc_valid': rfc_valid,
            'data_consistent': data_consistent,
            'curp_data': ai_result.get('curp_validation', {}),
            'rfc_data': ai_result.get('rfc_validation', {}),
            'extracted_fields': ai_result.get('ocr', {}).get('fields', {}),
            'errors': ai_result.get('errors', []),
            'can_auto_approve': can_auto_approve,
            'ai_result': ai_result
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': f'Error en validación IA: {str(e)}',
            'can_auto_approve': False
        }


def home(request):
    # Si no está logueado, mostrar landing page
    if not request.user.is_authenticated:
        return render(request, "documents/landing.html")
    
    # Si está logueado como admin, redirigir al panel admin
    if hasattr(request.user, 'profile') and request.user.profile.role == 'admin':
        return redirect('admin_panel')
    
    # Si es staff/senior, redirigir a su panel correspondiente
    if hasattr(request.user, 'profile'):
        if request.user.profile.role == 'validator1':
            return redirect('validator1_panel')
        elif request.user.profile.role == 'validator2':
            return redirect('validator2_panel')
    
    # Para clientes y otros usuarios
    if request.method == "POST":
        form = DocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                document = form.save(commit=False)
                document.user = request.user
                document.title = f"Solicitud {document.get_credit_type_display()}"
                document.save()
                
                # Extraer datos reales con OCR de todos los archivos subidos
                files_to_scan = []
                if document.acta_nacimiento:
                    files_to_scan.append(document.acta_nacimiento)
                if document.comprobante_domicilio:
                    files_to_scan.append(document.comprobante_domicilio)
                if document.ine:
                    files_to_scan.append(document.ine)
                if document.comprobante_bancario:
                    files_to_scan.append(document.comprobante_bancario)
                if document.constancia_fiscal:
                    files_to_scan.append(document.constancia_fiscal)
                if document.curp_documento:
                    files_to_scan.append(document.curp_documento)
                if document.file:
                    files_to_scan.append(document.file)
                
                ocr_fields = {}
                for f in files_to_scan:
                    try:
                        data = extract_text_from_pdf(str(f.path))
                        for key in ["nombre", "curp", "rfc", "domicilio", "telefono", "ocupacion", "estado_civil"]:
                            if data.get(key) and not ocr_fields.get(key):
                                ocr_fields[key] = data[key]
                    except Exception as e:
                        print(f"Error OCR en {f.name}: {e}")
                
                if ocr_fields:
                    if ocr_fields.get("nombre") and not document.nombre_completo:
                        document.nombre_completo = ocr_fields["nombre"]
                    if ocr_fields.get("curp"):
                        document.ocr_curp = ocr_fields["curp"]
                        if not document.curp:
                            document.curp = ocr_fields["curp"]
                    if ocr_fields.get("rfc"):
                        document.ocr_rfc = ocr_fields["rfc"]
                        if not document.rfc:
                            document.rfc = ocr_fields["rfc"]
                    if ocr_fields.get("domicilio") and not document.domicilio:
                        document.domicilio = ocr_fields["domicilio"]
                    if ocr_fields.get("telefono") and not document.telefono:
                        document.telefono = ocr_fields["telefono"]
                    document.save()
                
                if ocr_fields:
                    messages.success(request, f"Documento subido. OCR extrajo: {', '.join(k for k, v in ocr_fields.items() if v)}")
                else:
                    messages.success(request, "Documento subido. (OCR no detecto datos adicionales)")
                
                document.assign_worker()
                add_history(document, "pending_validator1", "Documento subido para validación.", request.user)
                messages.success(request, "Documento subido correctamente. Datos extraídos (OCR).")
                return redirect("home")
            except Exception:
                messages.error(request, "No se pudo subir el documento. Intenta de nuevo o contacta al administrador.")
    else:
        form = DocumentUploadForm()

    documents = Document.objects.filter(user=request.user).order_by("-uploaded_at")
    return render(request, "documents/home.html", {"form": form, "documents": documents})


@login_required
def click_continuidad(request, pk):
    document = get_object_or_404(Document, pk=pk, user=request.user)
    if document.status == "approved" and not document.clicked_continuidad:
        document.clicked_continuidad = timezone.now()
        document.save()
    return redirect("home")


@login_required
@user_passes_test(is_worker)
def worker_panel(request):
    documents = Document.objects.filter(status="pending", assigned_to=request.user).order_by("-uploaded_at")
    return render(request, "documents/worker_panel.html", {"documents": documents})


@login_required
def document_detail(request, pk):
    document = get_object_or_404(Document, pk=pk)
    if request.user != document.user and request.user != document.assigned_to:
        raise PermissionDenied

    comment_form = None
    correction_form = None
    if request.user == document.assigned_to and document.status == "pending":
        comment_form = DocumentCommentForm()
    if request.user == document.user and document.status == "rejected":
        correction_form = DocumentCorrectionForm(instance=document)

    if request.method == "POST":
        if "approve" in request.POST and request.user == document.assigned_to:
            try:
                document.status = "approved"
                document.save()
                add_history(document, "approved", "Documento aprobado.", request.user)
                messages.success(request, "Documento aprobado.")
                return redirect(reverse("worker_panel"))
            except Exception:
                messages.error(request, "No se pudo aprobar el documento. Intenta de nuevo.")

        if "reject" in request.POST and request.user == document.assigned_to:
            comment_form = DocumentCommentForm(request.POST)
            if comment_form.is_valid():
                try:
                    document.status = "rejected"
                    document.save()
                    comment = comment_form.save(commit=False)
                    comment.document = document
                    comment.author = request.user
                    comment.save()
                    add_history(document, "rejected", comment.comment, request.user)
                    messages.warning(request, "Documento rechazado y comentario guardado.")
                    return redirect(reverse("worker_panel"))
                except Exception:
                    messages.error(request, "No se pudo rechazar el documento. Intenta de nuevo.")

        if "resubmit" in request.POST and request.user == document.user and document.status == "rejected":
            correction_form = DocumentCorrectionForm(request.POST, request.FILES, instance=document)
            if correction_form.is_valid():
                try:
                    document = correction_form.save(commit=False)
                    document.status = "pending"
                    document.assign_worker()
                    document.save()
                    add_history(document, "pending", "Documento corregido y reenviado.", request.user)
                    messages.success(request, "Documento reenviado para revisión.")
                    return redirect("home")
                except Exception:
                    messages.error(request, "No se pudo reenviar el documento. Intenta de nuevo.")

    return render(request, "documents/document_detail.html", {
        "document": document,
        "comment_form": comment_form,
        "correction_form": correction_form,
    })


@login_required
@user_passes_test(is_worker)
def validator1_panel(request):
    """Panel para STAFF: revisa documentos pendientes con asignacion equitativa"""
    from django.db.models import Count
    
    # Obtener todos los STAFF activos
    all_staff = list(User.objects.filter(profile__role="validator1", profile__is_worker=True))
    
    # Contar documentos pendientes de cada STAFF
    counts = (
        Document.objects
        .filter(status="validator1_review", validator1__in=all_staff)
        .values('validator1')
        .annotate(count=Count('id'))
        .values_list('validator1', 'count')
    )
    count_map = dict(counts)
    
    # Documentos pendientes sin asignar
    unassigned = Document.objects.filter(
        status="pending_validator1",
        validator1__isnull=True
    ).order_by("uploaded_at")
    
    # Asignar equitativamente: darle al STAFF con menos carga
    for doc in unassigned:
        if all_staff:
            least_loaded = min(all_staff, key=lambda u: count_map.get(u.id, 0))
            doc.validator1 = least_loaded
            doc.status = "validator1_review"
            doc.save()
            count_map[least_loaded.id] = count_map.get(least_loaded.id, 0) + 1
    
    # Mostrar documentos asignados al usuario actual
    documents = Document.objects.filter(
        validator1=request.user,
        status="validator1_review"
    ).order_by("-uploaded_at")
    
    # Estadisticas
    pending_count = Document.objects.filter(
        status="validator1_review"
    ).count()
    reviewing_count = documents.count()
    approved_count = Document.objects.filter(
        validator1=request.user,
        validator1_approved=True
    ).count()
    rejected_count = Document.objects.filter(
        validator1=request.user,
        validator1_approved=False
    ).count()
    
    return render(request, "documents/validator1_dashboard.html", {
        "documents": documents,
        "pending_count": pending_count,
        "reviewing_count": reviewing_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
    })


def get_file_type(file_obj):
    """Detecta el tipo de archivo: 'pdf', 'image' o 'unknown'"""
    if not file_obj:
        return None
    file_name = file_obj.name.lower()
    if file_name.endswith(('.pdf',)):
        return "pdf"
    elif file_name.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
        return "image"
    return "unknown"


@login_required
@user_passes_test(is_worker)
def validator1_review(request, pk):
    """STAFF revisa un documento"""
    document = get_object_or_404(Document, pk=pk)
    
    # Asignar al validador si el documento aun esta pendiente
    if document.status == "pending_validator1" and not document.validator1:
        document.validator1 = request.user
        document.status = "validator1_review"
        document.save()
    
    if document.validator1 != request.user:
        raise PermissionDenied
    
    # Realiza validación automática (IA)
    ai_validation = perform_ai_validation(document)
    
    # Determina tipo de cliente (nuevo, antiguo con match, antiguo sin match)
    client_info = determine_client_type(document)
    
    # Detecta tipos de archivo para cada documento
    file_types = {
        'file': get_file_type(document.file),
        'acta_nacimiento': get_file_type(document.acta_nacimiento),
        'comprobante_domicilio': get_file_type(document.comprobante_domicilio),
        'ine': get_file_type(document.ine),
        'comprobante_bancario': get_file_type(document.comprobante_bancario),
        'constancia_fiscal': get_file_type(document.constancia_fiscal),
        'curp_documento': get_file_type(document.curp_documento),
    }
    
    if request.method == "POST":
        if "approve" in request.POST:
            document.validator1_approved = True
            document.validator1_comment = request.POST.get("comment", "")
            document.validator1_date = timezone.now()
            
            # Usa validaciones de IA si están disponibles
            if ai_validation.get('success'):
                # IA funcionó - validar según los resultados de IA
                if not ai_validation.get('curp_valid'):
                    messages.error(request, "CURP inválido detectado por IA. Por favor revisa el documento.")
                    return render(request, "documents/validator1_review.html", {
                        "document": document,
                        "ai_validation": ai_validation,
                        "file_types": file_types,
                        "client_info": client_info
                    })
                
                if not ai_validation.get('rfc_valid'):
                    messages.error(request, "RFC inválido detectado por IA. Por favor revisa el documento.")
                    return render(request, "documents/validator1_review.html", {
                        "document": document,
                        "ai_validation": ai_validation,
                        "file_types": file_types,
                        "client_info": client_info
                    })
            else:
                # IA no funcionó - el usuario debe revisar manualmente
                messages.info(request, "⚠️ Análisis automático no disponible. Revisa los documentos manualmente.")
            
            document.status = "pending_validator2"
            document.save()
            add_history(document, "validator1_approved", "Validación STAFF aprobada", request.user)
            messages.success(request, "Documento validado y enviado a SENIOR.")
            return redirect("validator1_panel")
        
        elif "reject" in request.POST:
            document.validator1_approved = False
            document.validator1_comment = request.POST.get("comment", "")
            document.validator1_date = timezone.now()
            document.rejection_reason = request.POST.get("reason", "otro")
            document.rejection_details = request.POST.get("details", "")
            document.status = "rejected"
            document.save()
            add_history(document, "validator1_rejected", f"Validación STAFF rechazada: {document.rejection_reason}", request.user)
            
            # Enviar correo de rechazo al cliente
            send_rejection_email(
                document.user.email,
                document.user.username,
                document.rejection_reason,
                document.rejection_details
            )
            
            messages.warning(request, "Documento rechazado y correo enviado al cliente.")
            return redirect("validator1_panel")
    
    return render(request, "documents/validator1_review.html", {
        "document": document,
        "ai_validation": ai_validation,
        "file_types": file_types,
        "client_info": client_info
    })


@login_required
@user_passes_test(is_worker)
def validator2_panel(request):
    """Panel para SENIOR: revisa documentos pendientes con asignacion equitativa"""
    from django.db.models import Count
    
    # Obtener todos los SENIOR activos
    all_senior = list(User.objects.filter(profile__role="validator2", profile__is_worker=True))
    
    # Contar documentos pendientes de cada SENIOR
    counts = (
        Document.objects
        .filter(status="validator2_review", validator2__in=all_senior)
        .values('validator2')
        .annotate(count=Count('id'))
        .values_list('validator2', 'count')
    )
    count_map = dict(counts)
    
    # Documentos pendientes sin asignar
    unassigned = Document.objects.filter(
        status="pending_validator2",
        validator2__isnull=True
    ).order_by("uploaded_at")
    
    # Asignar equitativamente: darle al SENIOR con menos carga
    for doc in unassigned:
        if all_senior:
            least_loaded = min(all_senior, key=lambda u: count_map.get(u.id, 0))
            doc.validator2 = least_loaded
            doc.status = "validator2_review"
            doc.save()
            count_map[least_loaded.id] = count_map.get(least_loaded.id, 0) + 1
    
    # Mostrar documentos asignados al usuario actual
    documents = Document.objects.filter(
        validator2=request.user,
        status="validator2_review"
    ).order_by("-uploaded_at")
    
    # Estadisticas
    pending_count = Document.objects.filter(
        status="validator2_review"
    ).count()
    reviewing_count = documents.count()
    approved_count = Document.objects.filter(
        validator2=request.user,
        validator2_approved=True
    ).count()
    rejected_count = Document.objects.filter(
        validator2=request.user,
        validator2_approved=False
    ).count()
    
    return render(request, "documents/validator2_dashboard.html", {
        "documents": documents,
        "pending_count": pending_count,
        "reviewing_count": reviewing_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
    })


@login_required
@user_passes_test(is_worker)
def validator2_review(request, pk):
    """SENIOR revisa un documento (confirmación final)"""
    document = get_object_or_404(Document, pk=pk)
    
    # Asignar al validador si el documento aun esta pendiente
    if document.status == "pending_validator2" and not document.validator2:
        document.validator2 = request.user
        document.status = "validator2_review"
        document.save()
    
    if document.validator2 != request.user:
        raise PermissionDenied
    
    # Realiza validación automática (IA) para mostrar datos extraídos
    ai_validation = perform_ai_validation(document)
    
    # Determina tipo de cliente (nuevo, antiguo con match, antiguo sin match)
    client_info = determine_client_type(document)
    
    # Detecta tipos de archivo para cada documento
    file_types = {
        'file': get_file_type(document.file),
        'acta_nacimiento': get_file_type(document.acta_nacimiento),
        'comprobante_domicilio': get_file_type(document.comprobante_domicilio),
        'ine': get_file_type(document.ine),
        'comprobante_bancario': get_file_type(document.comprobante_bancario),
        'constancia_fiscal': get_file_type(document.constancia_fiscal),
        'curp_documento': get_file_type(document.curp_documento),
    }
    
    if request.method == "POST":
        if "approve" in request.POST:
            document.validator2_approved = True
            document.validator2_comment = request.POST.get("comment", "")
            document.validator2_date = timezone.now()
            document.status = "approved"
            document.save()
            add_history(document, "validator2_approved", "Validación SENIOR aprobada - Solicitud Completa", request.user)
            
            # Obtener motivo de aprobación
            approval_reason = request.POST.get("approval_reason", "Documento aprobado por el equipo de validación")
            
            # Enviar correo de aprobación al cliente (INDEPENDIENTE del ERP)
            send_approval_email(
                document.user.email,
                document.user.username,
                document.get_credit_type_display(),
                approval_reason
            )
            
            # Intentar exportar a ERP (opcional - si falla no afecta el flujo)
            export_success = export_to_erp(document)
            if export_success:
                messages.success(request, "✅ Documento aprobado. Correo enviado y datos exportados a ERP.")
            else:
                messages.success(request, "✅ Documento aprobado y correo enviado. (Nota: ERP no disponible)")
            
            return redirect("validator2_panel")
        
        elif "reject" in request.POST:
            document.validator2_approved = False
            document.validator2_comment = request.POST.get("comment", "")
            document.validator2_date = timezone.now()
            document.rejection_reason = request.POST.get("reason", "otro")
            document.rejection_details = request.POST.get("details", "")
            document.status = "rejected"
            document.save()
            add_history(document, "validator2_rejected", f"Validación SENIOR rechazada: {document.rejection_reason}", request.user)
            
            # Enviar correo de rechazo al cliente
            send_rejection_email(
                document.user.email,
                document.user.username,
                document.rejection_reason,
                document.rejection_details
            )
            
            messages.warning(request, "Documento rechazado y correo enviado al cliente.")
            return redirect("validator2_panel")
    
    return render(request, "documents/validator2_review.html", {
        "document": document,
        "ai_validation": ai_validation,
        "file_types": file_types,
        "client_info": client_info
    })


def export_to_erp(document):
    """Exporta datos aprobados a ERP (simulado)
    Retorna True si es exitoso, False si falla
    El correo ya fue enviado antes de esta función
    """
    try:
        # Crear registro en tabla ERP
        erp_record = ERPExport.objects.create(
            document=document,
            nombre_completo=document.nombre_completo,
            curp=document.curp,
            rfc=document.rfc,
            domicilio=document.domicilio,
            telefono=document.telefono,
            credit_type=document.get_credit_type_display(),
            exported_by=document.validator2,
        )
        
        document.erp_exported = True
        document.erp_export_date = timezone.now()
        document.save()
        add_history(document, "erp_exported", "Datos exportados a ERP", None)
        
        print(f"✅ Datos exportados a ERP para documento #{document.id}")
        return True
    except Exception as e:
        print(f"⚠️ Advertencia: Error exportando a ERP: {e}")
        print(f"   El cliente recibió su correo de aprobación, pero ERP no está disponible")
        return False


# ==================== VISTAS DE ADMINISTRADOR ====================

@login_required
@user_passes_test(is_admin)
def admin_panel(request):
    """Dashboard principal del administrador - crear usuarios y ver estadísticas"""
    
    # Procesar creación de usuario si es POST
    if request.method == "POST" and "create_validator" in request.POST:
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        validator_type = request.POST.get("validator_type")
        first_name = request.POST.get("first_name", "Usuario").strip()
        last_name = request.POST.get("last_name", "").strip()
        
        if not username or not email or not validator_type:
            messages.error(request, "⚠️ Completa todos los campos requeridos.")
        elif User.objects.filter(username=username).exists():
            messages.error(request, f"❌ El usuario '{username}' ya existe.")
        elif User.objects.filter(email=email).exists():
            messages.error(request, f"❌ El email '{email}' ya está registrado.")
        else:
            # Crear usuario
            temp_password = "TempPass123!"
            user = User.objects.create_user(
                username=username,
                email=email,
                password=temp_password,
                first_name=first_name,
                last_name=last_name
            )
            user.profile.role = validator_type
            user.profile.is_worker = True
            user.profile.save()
            
            messages.success(
                request,
                f"✅ Usuario '{username}' creado. Contraseña temporal: {temp_password}"
            )
            
            # Log en auditoría
            AuditLog.objects.create(
                table_name="auth_user",
                row_pk=user.id,
                action="CREATE",
                changed_by=request.user,
                note=f"Usuario creado: {username} ({validator_type})"
            )
            
            return redirect("admin_panel")
    
    # Estadísticas generales
    total_documents = Document.objects.count()
    pending_documents = Document.objects.filter(status__in=["pending_validator1", "validator1_review", "pending_validator2", "validator2_review"]).count()
    approved_documents = Document.objects.filter(status="approved").count()
    rejected_documents = Document.objects.filter(status="rejected").count()
    erp_exported = Document.objects.filter(erp_exported=True).count()
    
    # Métrica de productividad STAFF
    # Meta: 154 solicitudes/día × 5 días = 770/semana por STAFF
    META_DIARIA = 154
    META_SEMANAL = 770
    HORAS_JORNADA = 9
    
    # Calcular inicio de la semana actual (lunes)
    today = timezone.now()
    day_of_week = today.weekday()
    start_of_week = today - timedelta(days=day_of_week)
    week_start_date = timezone.make_aware(datetime(start_of_week.year, start_of_week.month, start_of_week.day))
    
    # Obtener todos los STAFF
    staff_users = User.objects.filter(profile__role="validator1")
    
    staff_metrics = []
    total_staff_processed = 0
    total_staff_meta = 0
    
    for staff in staff_users:
        # Documentos procesados esta semana por este STAFF
        weekly_processed = Document.objects.filter(
            validator1=staff,
            status__in=["validator1_review", "approved", "rejected"],
            validator1_date__gte=week_start_date
        ).count()
        
        # Promedio diario (dividir entre días laborales transcurridos o 5)
        days_elapsed = min(day_of_week + 1, 5)
        daily_avg = weekly_processed / days_elapsed if days_elapsed > 0 else 0
        
        meets_daily = daily_avg >= META_DIARIA
        meets_weekly = weekly_processed >= META_SEMANAL
        
        total_staff_processed += weekly_processed
        total_staff_meta += META_SEMANAL
        
        staff_metrics.append({
            'user': staff,
            'weekly_processed': weekly_processed,
            'daily_avg': round(daily_avg, 1),
            'meets_daily': meets_daily,
            'meets_weekly': meets_weekly,
            'progress_pct': min(round((weekly_processed / META_SEMANAL) * 100, 1), 100),
        })
    
    # Resumen general del equipo STAFF
    team_progress_pct = round((total_staff_processed / total_staff_meta) * 100, 1) if total_staff_meta > 0 else 0
    team_meets = team_progress_pct >= 100
    
    # Metricas de productividad SENIOR
    senior_users = User.objects.filter(profile__role="validator2")
    senior_metrics = []
    total_senior_approved = 0
    total_senior_rejected = 0
    
    for senior in senior_users:
        weekly_approved = Document.objects.filter(
            validator2=senior,
            validator2_approved=True,
            validator2_date__gte=week_start_date
        ).count()
        
        weekly_rejected = Document.objects.filter(
            validator2=senior,
            validator2_approved=False,
            validator2_date__gte=week_start_date
        ).count()
        
        weekly_total = weekly_approved + weekly_rejected
        
        total_senior_approved += weekly_approved
        total_senior_rejected += weekly_rejected
        
        senior_metrics.append({
            'user': senior,
            'weekly_approved': weekly_approved,
            'weekly_rejected': weekly_rejected,
            'weekly_total': weekly_total,
            'approval_rate': round((weekly_approved / weekly_total) * 100, 1) if weekly_total > 0 else 0,
        })
    
    # Estadísticas por tipo de crédito
    credito_total = Document.objects.filter(credit_type="credito").count()
    tarjeta_total = Document.objects.filter(credit_type="tarjeta").count()
    
    # Métrica de CAPTACION DE CLIENTES (cliente bueno vs perdido)
    # Captacion de clientes: aprobados que dieron clic en "Ver línea de crédito"
    captacion_clientes = Document.objects.filter(
        status="approved",
        clicked_continuidad__isnull=False
    ).count()
    
    # Cliente perdido: aprobados que NO dieron clic y pasaron 5+ días desde que crearon la solicitud
    five_days_ago = timezone.now() - timedelta(days=5)
    continuidad_perdidos = Document.objects.filter(
        status="approved",
        clicked_continuidad__isnull=True,
        uploaded_at__lte=five_days_ago
    ).count()
    
    # Aprobados sin clic (aún dentro del plazo de 5 días)
    aprobados_pendientes_clic = Document.objects.filter(
        status="approved",
        clicked_continuidad__isnull=True,
        uploaded_at__gt=five_days_ago
    ).count()
    
    # Tasa de continuidad
    total_resueltos = captacion_clientes + continuidad_perdidos
    tasa_continuidad = round((captacion_clientes / total_resueltos) * 100, 1) if total_resueltos > 0 else 0
    
    # Información de usuarios
    validator1_users = User.objects.filter(profile__role="validator1")
    validator2_users = User.objects.filter(profile__role="validator2")
    client_users = User.objects.filter(profile__role="cliente")
    
    # Usuarios con estadísticas
    validators_stats = []
    for validator in validator1_users | validator2_users:
        role = validator.profile.role
        if role == "validator1":
            approved = Document.objects.filter(validator1=validator, validator1_approved=True).count()
            rejected = Document.objects.filter(validator1=validator, validator1_approved=False).count()
        else:
            approved = Document.objects.filter(validator2=validator, validator2_approved=True).count()
            rejected = Document.objects.filter(validator2=validator, validator2_approved=False).count()
        
        validators_stats.append({
            'user': validator,
            'role': validator.profile.get_role_display(),
            'approved': approved,
            'rejected': rejected,
        })
    
    return render(request, "documents/admin_dashboard.html", {
        "total_documents": total_documents,
        "pending_documents": pending_documents,
        "approved_documents": approved_documents,
        "rejected_documents": rejected_documents,
        "erp_exported": erp_exported,
        "credito_total": credito_total,
        "tarjeta_total": tarjeta_total,
        "validator1_users": validator1_users,
        "validator2_users": validator2_users,
        "client_users": client_users,
        "validators_stats": validators_stats,
        "staff_metrics": staff_metrics,
        "team_progress_pct": team_progress_pct,
        "team_meets": team_meets,
        "meta_diaria": META_DIARIA,
        "meta_semanal": META_SEMANAL,
        "horas_jornada": HORAS_JORNADA,
        "total_staff_processed": total_staff_processed,
        "total_staff_meta": total_staff_meta,
        "senior_metrics": senior_metrics,
        "total_senior_approved": total_senior_approved,
        "total_senior_rejected": total_senior_rejected,
        "captacion_clientes": captacion_clientes,
        "continuidad_perdidos": continuidad_perdidos,
        "aprobados_pendientes_clic": aprobados_pendientes_clic,
        "tasa_continuidad": tasa_continuidad,
    })



@login_required
@user_passes_test(is_admin)
def admin_audit_log(request):
    """Ver registro de auditoría - todos los cambios en el sistema"""
    
    # Obtener filtros
    filter_action = request.GET.get("action", "")
    filter_user = request.GET.get("user", "")
    filter_table = request.GET.get("table", "")
    
    # Iniciar queryset
    logs = AuditLog.objects.select_related("changed_by").order_by("-changed_at")
    
    # Aplicar filtros
    if filter_action:
        logs = logs.filter(action=filter_action)
    if filter_user:
        logs = logs.filter(changed_by_id=filter_user)
    if filter_table:
        logs = logs.filter(table_name=filter_table)
    
    # Obtener opciones para filtros
    available_actions = AuditLog.objects.values_list("action", flat=True).distinct()
    available_users = User.objects.filter(audit_logs__isnull=False).distinct()
    available_tables = AuditLog.objects.values_list("table_name", flat=True).distinct()
    
    # Contar antes de paginar
    logs_count = logs.count()
    
    # Paginación simple (últimos 100 registros)
    logs_list = list(logs[:100])
    
    # Obtener cambios en documentos para información adicional
    document_history = DocumentHistory.objects.select_related("document", "author").order_by("-created_at")[:100]
    document_history_count = len(document_history)
    
    return render(request, "documents/admin_audit_log.html", {
        "logs": logs_list,
        "logs_count": logs_count,
        "document_history": document_history,
        "document_history_count": document_history_count,
        "filter_action": filter_action,
        "filter_user": filter_user,
        "filter_table": filter_table,
        "available_actions": sorted(set(available_actions)),
        "available_users": available_users,
        "available_tables": sorted(set(available_tables)),
    })


@login_required
@user_passes_test(is_admin)
def admin_edit_validator(request, pk):
    """Editar información de un usuario"""
    try:
        validator_user = User.objects.get(pk=pk, profile__role__in=['validator1', 'validator2'])
    except User.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('admin_panel')

    if request.method == "POST":
        # Obtener datos del formulario
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip()
        validator_type = request.POST.get("validator_type", "").strip()
        
        # Validar datos
        errors = []
        
        if not first_name:
            errors.append("El nombre es requerido.")
        if not last_name:
            errors.append("El apellido es requerido.")
        if not email:
            errors.append("El email es requerido.")
        if not validator_type or validator_type not in ['validator1', 'validator2']:
            errors.append("Tipo de rol inválido.")
        
        # Verificar email único (excepto para el usuario actual)
        if email != validator_user.email:
            if User.objects.filter(email=email).exists():
                errors.append(f"El email '{email}' ya está registrado.")
        
        if errors:
            return render(request, "documents/admin_edit_validator.html", {
                "validator": validator_user,
                "errors": errors
            })
        
        # Actualizar usuario
        validator_user.first_name = first_name
        validator_user.last_name = last_name
        validator_user.email = email
        validator_user.profile.role = validator_type
        validator_user.profile.save()
        validator_user.save()
        
        # Registrar en auditoría
        AuditLog.objects.create(
            table_name="auth_user",
            row_pk=validator_user.id,
            action="UPDATE",
            changed_by=request.user,
            note=f"Datos de usuario actualizados: {first_name} {last_name} ({validator_type})"
        )
        
        messages.success(request, f"✅ Usuario '{validator_user.username}' actualizado exitosamente.")
        return redirect('admin_panel')
    
    return render(request, "documents/admin_edit_validator.html", {
        "validator": validator_user
    })


@login_required
@user_passes_test(is_admin)
def admin_delete_validator(request, pk):
    """Eliminar un usuario del sistema"""
    try:
        validator_user = User.objects.get(pk=pk, profile__role__in=['validator1', 'validator2'])
    except User.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
        return redirect('admin_panel')

    # Verificar que no haya documentos asignados
    has_pending_docs = Document.objects.filter(
        validator1=validator_user,
        status__in=['validator1_review', 'pending_validator1']
    ).exists() or Document.objects.filter(
        validator2=validator_user,
        status__in=['validator2_review', 'pending_validator2']
    ).exists()
    
    if request.method == "POST":
        confirm = request.POST.get("confirm", "").lower()
        
        if confirm != "si":
            messages.error(request, "Debe confirmar la eliminación.")
            return render(request, "documents/admin_delete_validator.html", {
                "validator": validator_user,
                "has_pending_docs": has_pending_docs
            })
        
        if has_pending_docs:
            messages.error(request, "No se puede eliminar un usuario con documentos pendientes.")
            return render(request, "documents/admin_delete_validator.html", {
                "validator": validator_user,
                "has_pending_docs": has_pending_docs
            })
        
        # Guardar información para auditoría
        username = validator_user.username
        email = validator_user.email
        
        # Registrar en auditoría antes de eliminar
        AuditLog.objects.create(
            table_name="auth_user",
            row_pk=validator_user.id,
            action="DELETE",
            changed_by=request.user,
            note=f"Usuario eliminado: {username} ({email})"
        )
        
        # Eliminar usuario
        validator_user.delete()
        
        messages.success(request, f"Usuario '{username}' eliminado exitosamente del sistema.")
        return redirect('admin_panel')
    
    return render(request, "documents/admin_delete_validator.html", {
        "validator": validator_user,
        "has_pending_docs": has_pending_docs
    })


@login_required
@user_passes_test(is_admin)
def download_approved_documents(request):
    """Descarga todos los documentos aprobados organizados por cliente como ZIP"""
    
    # Obtener todos los documentos aprobados
    approved_docs = Document.objects.filter(status='approved').select_related('user')
    
    if not approved_docs.exists():
        messages.warning(request, "No hay documentos aprobados para descargar.")
        return redirect('admin_panel')
    
    # Crear ZIP en memoria
    zip_buffer = io.BytesIO()
    
    try:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Agrupar por cliente
            clients_docs = {}
            for doc in approved_docs:
                client_name = f"{doc.user.username}_{doc.user.id}"
                if client_name not in clients_docs:
                    clients_docs[client_name] = {
                        'user': doc.user,
                        'documents': []
                    }
                clients_docs[client_name]['documents'].append(doc)
            
            # Crear estructura por cliente
            for client_folder, client_data in clients_docs.items():
                user = client_data['user']
                documents = client_data['documents']
                
                # Crear carpeta del cliente
                client_path = f"Clientes/{client_folder}"
                
                # 1. Crear archivo de información del cliente
                client_info = f"""INFORMACIÓN DEL CLIENTE
================================

Nombre de usuario: {user.username}
Email: {user.email}
Nombre completo: {user.get_full_name() if user.get_full_name() else 'N/A'}
Fecha de registro: {user.date_joined.strftime('%d/%m/%Y %H:%M')}
ID Usuario: {user.id}

DOCUMENTOS APROBADOS
================================
Total de documentos: {len(documents)}

"""
                
                # Agregar info de cada documento
                for idx, doc in enumerate(documents, 1):
                    client_info += f"""
--- DOCUMENTO {idx} ---
Título: {doc.title}
Tipo: {doc.document_type if hasattr(doc, 'document_type') else 'N/A'}
Fecha de subida: {doc.uploaded_at.strftime('%d/%m/%Y %H:%M')}
Estado: {doc.get_status_display()}
STAFF: {doc.validator1.username if doc.validator1 else 'Sin asignar'}
SENIOR: {doc.validator2.username if doc.validator2 else 'Sin asignar'}
Fecha de aprobación: {doc.validator2_date.strftime('%d/%m/%Y %H:%M') if doc.validator2_date else 'N/A'}

"""
                    
                    # Agregar datos extraídos si existen
                    if doc.curp or doc.rfc or doc.nombre_completo:
                        client_info += """Datos Extraídos por IA:
"""
                        if doc.curp:
                            client_info += f"  • CURP: {doc.curp}\n"
                        if doc.rfc:
                            client_info += f"  • RFC: {doc.rfc}\n"
                        if doc.nombre_completo:
                            client_info += f"  • Nombre: {doc.nombre_completo}\n"
                        if doc.domicilio:
                            client_info += f"  • Dirección: {doc.domicilio}\n"
                        if doc.telefono:
                            client_info += f"  • Teléfono: {doc.telefono}\n"
                        client_info += "\n"
                
                # Agregar archivo de información al ZIP
                zip_file.writestr(
                    f"{client_path}/00_INFORMACIÓN_CLIENTE.txt",
                    client_info.encode('utf-8')
                )
                
                # 2. Agregar documentos originales
                for idx, doc in enumerate(documents, 1):
                    # Obtener archivo del documento
                    if doc.file and os.path.exists(doc.file.path):
                        with open(doc.file.path, 'rb') as f:
                            # Obtener extensión del archivo
                            original_filename = os.path.basename(doc.file.name)
                            zip_path = f"{client_path}/Documentos/{idx:02d}_{doc.title}_{original_filename}"
                            zip_file.writestr(zip_path, f.read())
                
                # 3. Crear CSV con datos extraídos
                csv_content = "Título,Tipo,CURP,RFC,Nombre,Teléfono,Email,Dirección,Ocupación,Estado Civil,Fecha Aprobación\n"
                for doc in documents:
                    csv_line = f'"{doc.title}","{doc.document_type if hasattr(doc, "document_type") else "N/A"}","{doc.curp or ""}","{doc.rfc or ""}","{doc.nombre_completo or ""}","{doc.telefono or ""}","{doc.email or ""}","{doc.domicilio or ""}","{doc.ocupacion or ""}","{doc.estado_civil or ""}","{doc.validator2_date.strftime("%d/%m/%Y") if doc.validator2_date else "N/A"}"\n'
                    csv_content += csv_line
                
                zip_file.writestr(
                    f"{client_path}/DATOS_EXTRAÍDOS.csv",
                    csv_content.encode('utf-8')
                )
        
        # Preparar respuesta HTTP
        zip_buffer.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Documentos_Aprobados_{timestamp}.zip"
        
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Registrar descarga en auditoría
        AuditLog.objects.create(
            table_name="documents_erp_export",
            row_pk=0,
            action="EXPORT",
            changed_by=request.user,
            note=f"Descarga de {approved_docs.count()} documentos aprobados en formato ZIP"
        )
        
        return response
    
    except Exception as e:
        messages.error(request, f"Error al generar descarga: {str(e)}")
        return redirect('admin_panel')


@login_required
@user_passes_test(is_admin)
def monthly_extraction(request):
    """Extrae información de documentos por mes con filtrado por estado"""
    
    # Obtener parámetros
    year = request.GET.get('year')
    month = request.GET.get('month')
    status_filter = request.GET.get('status', '')
    day_filter = request.GET.get('day', '')
    
    # Obtener años disponibles en documentos
    years_available = list(set(
        doc.uploaded_at.year for doc in Document.objects.all() if doc.uploaded_at
    ))
    years_available.sort(reverse=True)
    
    # Si no se especifica año/mes, usar mes actual
    today = timezone.now()
    if not year:
        year = str(today.year)
    if not month:
        month = str(today.month).zfill(2)
    
    # Convertir a enteros
    try:
        year = int(year)
        month = int(month)
    except (ValueError, TypeError):
        year = today.year
        month = today.month
    
    # Validar rango
    if month < 1 or month > 12:
        month = today.month
    
    # Filtrar documentos por mes
    start_date = timezone.make_aware(datetime(year, month, 1))
    if month == 12:
        end_date = timezone.make_aware(datetime(year + 1, 1, 1))
    else:
        end_date = timezone.make_aware(datetime(year, month + 1, 1))
    
    documents = Document.objects.filter(
        uploaded_at__gte=start_date,
        uploaded_at__lt=end_date
    ).order_by('-uploaded_at')
    
    # Aplicar filtro de estado si se especifica
    if status_filter:
        documents = documents.filter(status=status_filter)
    
    # Aplicar filtro de día específico si se especifica
    if day_filter:
        try:
            day_date = timezone.make_aware(datetime.strptime(day_filter, '%Y-%m-%d'))
            next_day = day_date + timedelta(days=1)
            documents = documents.filter(
                uploaded_at__gte=day_date,
                uploaded_at__lt=next_day
            )
        except (ValueError, TypeError):
            pass
    
    # Contar por estado
    status_counts = {
        'pending_validator1': documents.filter(status='pending_validator1').count(),
        'validator1_review': documents.filter(status='validator1_review').count(),
        'pending_validator2': documents.filter(status='pending_validator2').count(),
        'validator2_review': documents.filter(status='validator2_review').count(),
        'approved': documents.filter(status='approved').count(),
        'rejected': documents.filter(status='rejected').count(),
    }
    
    # Conteo total
    total = documents.count()
    
    return render(request, 'documents/monthly_extraction.html', {
        'documents': documents,
        'year': year,
        'month': month,
        'month_name': ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                       'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][month],
        'status_filter': status_filter,
        'day_filter': day_filter,
        'status_counts': status_counts,
        'total': total,
        'years_available': years_available,
        'all_statuses': [
            ('validator1_review', 'En revisión STAFF'),
            ('validator2_review', 'En revisión SENIOR'),
            ('approved', 'Aprobado'),
            ('rejected', 'Rechazado'),
        ],
    })


@login_required
@user_passes_test(is_admin)
def export_monthly_excel(request):
    """Exporta datos del corte mensual a archivo Excel"""
    year = request.GET.get('year')
    month = request.GET.get('month')
    status_filter = request.GET.get('status', '')
    day_filter = request.GET.get('day', '')
    
    today = timezone.now()
    try:
        year = int(year) if year else today.year
        month = int(month) if month else today.month
    except (ValueError, TypeError):
        year = today.year
        month = today.month
    
    start_date = timezone.make_aware(datetime(year, month, 1))
    if month == 12:
        end_date = timezone.make_aware(datetime(year + 1, 1, 1))
    else:
        end_date = timezone.make_aware(datetime(year, month + 1, 1))
    
    documents = Document.objects.filter(
        uploaded_at__gte=start_date,
        uploaded_at__lt=end_date
    ).select_related('user', 'validator1', 'validator2').order_by('-uploaded_at')
    
    if status_filter:
        documents = documents.filter(status=status_filter)
    
    if day_filter:
        try:
            day_date = timezone.make_aware(datetime.strptime(day_filter, '%Y-%m-%d'))
            next_day = day_date + timedelta(days=1)
            documents = documents.filter(
                uploaded_at__gte=day_date,
                uploaded_at__lt=next_day
            )
        except (ValueError, TypeError):
            pass
    
    status_counts = {
        'pending_validator1': documents.filter(status='pending_validator1').count(),
        'validator1_review': documents.filter(status='validator1_review').count(),
        'pending_validator2': documents.filter(status='pending_validator2').count(),
        'validator2_review': documents.filter(status='validator2_review').count(),
        'approved': documents.filter(status='approved').count(),
        'rejected': documents.filter(status='rejected').count(),
    }
    
    month_name = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                  'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][month]
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"
    
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='111827')
    normal_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = f"Corte Mensual - {month_name} {year}"
    title_cell.font = title_font
    
    ws.merge_cells('A2:N2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Total de documentos: {documents.count()}"
    subtitle_cell.font = Font(name='Calibri', size=11, italic=True, color='6b7280')
    
    headers = [
        'ID', 'Cliente', 'Email', 'Título', 'Tipo de Crédito', 'Estado',
        'CURP', 'RFC', 'Nombre Completo', 'Teléfono', 'Domicilio',
        'STAFF', 'SENIOR', 'Fecha de Subida'
    ]
    
    col_widths = [6, 20, 30, 25, 18, 22, 20, 15, 30, 15, 35, 20, 20, 20]
    
    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    status_map = {
        'validator1_review': 'En revisión STAFF',
        'validator2_review': 'En revisión SENIOR',
        'approved': 'Aprobado',
        'rejected': 'Rechazado',
    }
    
    for row_num, doc in enumerate(documents, 5):
        data = [
            doc.id,
            doc.user.username,
            doc.user.email,
            doc.title,
            doc.get_credit_type_display(),
            status_map.get(doc.status, doc.status),
            doc.curp or '',
            doc.rfc or '',
            doc.nombre_completo or '',
            doc.telefono or '',
            doc.domicilio or '',
            doc.validator1.username if doc.validator1 else '',
            doc.validator2.username if doc.validator2 else '',
            doc.uploaded_at.strftime('%d/%m/%Y %H:%M') if doc.uploaded_at else '',
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            
            if doc.status == 'approved':
                cell.fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
            elif doc.status == 'rejected':
                cell.fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
    
    summary_row = row_num + 2 if documents.exists() else 6
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    ws.cell(row=summary_row, column=1, value='Resumen por Estado:').font = Font(name='Calibri', bold=True, size=12)
    
    summary_data = [
        ('En revisión STAFF', status_counts.get('validator1_review', 0)),
        ('En revisión SENIOR', status_counts.get('validator2_review', 0)),
        ('Aprobados', status_counts.get('approved', 0)),
        ('Rechazados', status_counts.get('rejected', 0)),
        ('TOTAL', documents.count()),
    ]
    
    for i, (label, count) in enumerate(summary_data):
        r = summary_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = Font(name='Calibri', bold=True if label == 'TOTAL' else False, size=11)
        ws.cell(row=r, column=2, value=count).font = Font(name='Calibri', bold=True if label == 'TOTAL' else False, size=11)
    
    filename = f"Corte_{month_name}_{year}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def delete_monthly_documents(request):
    """Elimina documentos del corte mensual con confirmación de contraseña"""
    
    if request.method == 'POST':
        password = request.POST.get('password', '')
        year = request.POST.get('year')
        month = request.POST.get('month')
        status_filter = request.POST.get('status_filter', '')
        day_filter = request.POST.get('day_filter', '')
        
        # Verificar contraseña
        if not request.user.check_password(password):
            messages.error(request, "Contraseña incorrecta. No se realizó ninguna acción.")
            return redirect('monthly_extraction')
        
        # Obtener fechas
        try:
            year = int(year)
            month = int(month)
        except (ValueError, TypeError):
            messages.error(request, "Parámetros inválidos.")
            return redirect('monthly_extraction')
        
        # Filtrar documentos por mes
        start_date = timezone.make_aware(datetime(year, month, 1))
        if month == 12:
            end_date = timezone.make_aware(datetime(year + 1, 1, 1))
        else:
            end_date = timezone.make_aware(datetime(year, month + 1, 1))
        
        documents = Document.objects.filter(
            uploaded_at__gte=start_date,
            uploaded_at__lt=end_date
        )
        
        # Aplicar filtro de estado
        if status_filter:
            documents = documents.filter(status=status_filter)
        
        # Aplicar filtro de día
        if day_filter:
            try:
                day_date = timezone.make_aware(datetime.strptime(day_filter, '%Y-%m-%d'))
                next_day = day_date + timedelta(days=1)
                documents = documents.filter(
                    uploaded_at__gte=day_date,
                    uploaded_at__lt=next_day
                )
            except (ValueError, TypeError):
                pass
        
        # Contar y eliminar
        count = documents.count()
        documents.delete()
        
        messages.success(request, f"✅ Se eliminaron {count} documento(s) permanentemente.")
        return redirect('monthly_extraction')
    
    return redirect('monthly_extraction')


